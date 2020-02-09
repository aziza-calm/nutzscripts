import openpyxl as xl
import pymorphy2
morph = pymorphy2.MorphAnalyzer()

wto = xl.load_workbook("to.xlsx")
wfrom = xl.load_workbook("from.xlsx")

wsto = wto['Лист1']
wsfrom = wfrom['Лист1']

for row in wsfrom.iter_rows(min_row=2):
    for cell_from in row:
        if cell_from.column_letter == 'D': # фамилия
            wsto.cell(row=cell_from.row + 2, column=1, value=cell_from.value)
        if cell_from.column_letter == 'E': # имя
            wsto.cell(row=cell_from.row + 2, column=2, value=cell_from.value)
        if cell_from.column_letter == 'F': # отчество
            wsto.cell(row=cell_from.row + 2, column=3, value=cell_from.value)
        if cell_from.column_letter == 'X': # дата рождения
            wsto.cell(row=cell_from.row + 2, column=4, value=cell_from.value)
        if cell_from.column_letter == 'V': # страна
            wsto.cell(row=cell_from.row + 2, column=6, value=cell_from.value)
        if cell_from.column_letter == 'W': # город
            wsto.cell(row=cell_from.row + 2, column=7, value=cell_from.value)
        if cell_from.column_letter == 'Q': # скайп
            wsto.cell(row=cell_from.row + 2, column=17, value=cell_from.value)
        if cell_from.column_letter == 'Y': # linkedin
            wsto.cell(row=cell_from.row + 2, column=18, value=cell_from.value)
        if cell_from.column_letter == 'Z': # facebook
            wsto.cell(row=cell_from.row + 2, column=19, value=cell_from.value)
        if cell_from.column_letter == 'AA': # vk
            wsto.cell(row=cell_from.row + 2, column=20, value=cell_from.value)
        if cell_from.column_letter == 'AB': # odnoklassniki
            wsto.cell(row=cell_from.row + 2, column=21, value=cell_from.value)
        if cell_from.column_letter == 'AC': # hyperlink
            wsto.cell(row=cell_from.row + 2, column=22, value=cell_from.value)
        if cell_from.column_letter == 'R': # work
            wsto.cell(row=cell_from.row + 2, column=23, value=cell_from.value)
        if cell_from.column_letter == 'T': # company sphere
            wsto.cell(row=cell_from.row + 2, column=25, value=cell_from.value)
        if cell_from.column_letter == 'S': # alumnus sphere
            wsto.cell(row=cell_from.row + 2, column=26, value=cell_from.value)
        if cell_from.column_letter == 'U': # industry
            wsto.cell(row=cell_from.row + 2, column=27, value=cell_from.value)
        if cell_from.column_letter == 'I':
            if str(cell_from.value).strip().endswith("mipt.ru"):
                wsto.cell(row=cell_from.row + 2, column=9, value=cell_from.value)
            elif str(cell_from.value).strip().endswith("phystech.edu"):
                wsto.cell(row=cell_from.row + 2, column=8, value=cell_from.value)
            else:
                wsto.cell(row=cell_from.row + 2, column=11, value=cell_from.value)
        if cell_from.column_letter == 'J':
            if str(cell_from.value).strip().endswith("mipt.ru"):
                wsto.cell(row=cell_from.row + 2, column=9, value=cell_from.value)
            elif str(cell_from.value).strip().endswith("phystech.edu"):
                wsto.cell(row=cell_from.row + 2, column=8, value=cell_from.value)
            else:
                wsto.cell(row=cell_from.row + 2, column=11, value=cell_from.value)
        if cell_from.column_letter == 'K':
            if str(cell_from.value).strip().endswith("mipt.ru"):
                wsto.cell(row=cell_from.row + 2, column=9, value=cell_from.value)
            elif str(cell_from.value).strip().endswith("phystech.edu"):
                wsto.cell(row=cell_from.row + 2, column=8, value=cell_from.value)
            else:
                wsto.cell(row=cell_from.row + 2, column=11, value=cell_from.value)
        if cell_from.column_letter == 'N': # telefonnummer
            wsto.cell(row=cell_from.row + 2, column=12, value=cell_from.value)
        if cell_from.column_letter == 'M': # telefonnummer extra
            wsto.cell(row=cell_from.row + 2, column=13, value=cell_from.value)
        if cell_from.column_letter == 'A': # post
            wsto.cell(row=cell_from.row + 2, column=24, value=cell_from.value)
        if cell_from.column_letter == 'E':
            parse_result = morph.parse(cell_from.value)
            if len(parse_result) > 0:
                if parse_result[0].tag.gender == 'masc':
                    wsto.cell(row=cell_from.row + 2, column=5, value="Мужской")
                if parse_result[0].tag.gender == 'femn':
                    wsto.cell(row=cell_from.row + 2, column=5, value="Женский")
        

wto.save("result.xlsx")